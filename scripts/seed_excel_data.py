#!/usr/bin/env python3
"""Extract ALL transactional data from Excel and seed into Neon PostgreSQL."""

import os
import sys
import uuid
import openpyxl
import psycopg2
from datetime import datetime, date

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "../../Nampark spreadsheet.xlsx")
DB_URL = os.environ["DATABASE_URL"]

def cuid():
    return "c" + uuid.uuid4().hex[:24]

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s in ("#VALUE!", "sunday"):
        return None
    for fmt in ["%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s.split(" ")[0], fmt).date()
        except (ValueError, IndexError):
            continue
    return None

def pf(val):
    try: return float(val) if val else 0.0
    except: return 0.0

def pi(val):
    try: return int(float(val)) if val else 0
    except: return 0

def normalize_return_type(val):
    s = str(val).strip().lower() if val else ""
    if "wrong" in s: return "WRONG_ITEM"
    if "missing" in s: return "MISSING_ITEM"
    if "cancel" in s: return "CANCELLED_ORDER"
    if "damage" in s: return "DAMAGED"
    if "expi" in s: return "EXPIRED"
    return "WRONG_ITEM"

driver_route_map = {
    "SAMUEL LUKAYO": ["MAJENGO", "MOLO"],
    "JOSEPH KAMAU": ["GAKOE", "MATAARA"],
    "SAMUEL MUCHAI": ["KANDARA", "NGARARIA"],
    "BRIAN MWANGI": ["GATUNDU", "MUNUNGA"],
}

driver_vehicle_reg = {
    "SAMUEL LUKAYO": "KDW 852B",
    "JOSEPH KAMAU": "KDN 396Q",
    "SAMUEL MUCHAI": "KDX 478B",
    "BRIAN MWANGI": "KDL 166M",
}

rep_route_map = {
    "NAHASHON NENE": ["GATUNDU"],
    "JOSEPH MACHARIA": ["MUNUNGA"],
    "KELVIN MWANGI": ["MOLO", "GAKOE"],
    "MATTHEW ROP": ["KANDARA"],
    "BERNARD RONO": ["MATAARA", "MAJENGO"],
    "ELIJAH KAMAU": ["NGARARIA"],
}

# Default rep for a route (when returns don't specify)
route_default_rep = {
    "GATUNDU": "NAHASHON NENE",
    "MUNUNGA": "JOSEPH MACHARIA",
    "MOLO": "KELVIN MWANGI",
    "GAKOE": "KELVIN MWANGI",
    "KANDARA": "MATTHEW ROP",
    "MATAARA": "BERNARD RONO",
    "MAJENGO": "BERNARD RONO",
    "NGARARIA": "ELIJAH KAMAU",
}


def main():
    print("Connecting to Neon...")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = True
    cur = conn.cursor()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Load lookups
    cur.execute("SELECT id, name FROM routes")
    route_map = {}
    route_map_upper = {}
    for rid, rname in cur.fetchall():
        route_map[rname.strip()] = rid
        route_map_upper[rname.strip().upper()] = rid

    cur.execute("SELECT id, name FROM sales_reps")
    rep_map = {rname.strip().upper(): rid for rid, rname in cur.fetchall()}

    cur.execute("SELECT id, name FROM drivers")
    driver_map = {rname.strip().upper(): rid for rid, rname in cur.fetchall()}

    cur.execute("SELECT id, name FROM sku_catalog")
    sku_map = {rname.strip().upper(): sid for sid, rname in cur.fetchall()}
    sku_list = list(sku_map.values())

    cur.execute("SELECT id, registration FROM vehicles")
    vehicle_map = {reg.strip(): vid for vid, reg in cur.fetchall()}

    def find_route_id(name):
        n = name.strip().upper()
        if n in route_map_upper:
            return route_map_upper[n]
        for rn, rid in route_map_upper.items():
            if n in rn or rn in n:
                return rid
        return None

    def find_rep_id(name):
        n = name.strip().upper()
        if n in rep_map:
            return rep_map[n]
        for rn, rid in rep_map.items():
            if n in rn or rn in n:
                return rid
        return list(rep_map.values())[0] if rep_map else None

    def find_driver_id(name):
        n = name.strip().upper()
        if n in driver_map:
            return driver_map[n]
        for rn, rid in driver_map.items():
            if n in rn or rn in n:
                return rid
        return None

    def find_sku_id(product):
        pu = str(product).strip().upper()
        if not pu:
            return sku_list[0] if sku_list else None
        for sn, sid in sku_map.items():
            words = set(sn.split())
            pwords = set(pu.split())
            if len(words & pwords) >= 1:
                return sid
        return sku_list[0] if sku_list else None

    def find_driver_for_route(route_name):
        rn = route_name.strip().upper()
        for dr_name, dr_routes in driver_route_map.items():
            if rn in [r.upper() for r in dr_routes]:
                return dr_name
        return None

    # ================================================================
    # PHASE 1: Collect ALL unique (date, route) pairs from ALL sheets
    # ================================================================
    print("\n=== Phase 1: Collecting all dates & routes ===")
    all_date_routes = set()  # (date, route_name_upper)

    # From Sales breakdown
    ws = wb["Sales breakdown"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date, week, rep, route = row[:4]
        if not route:
            continue
        d = parse_date(raw_date)
        if d:
            all_date_routes.add((d, str(route).strip().upper()))

    # From Returns
    ws = wb["Returns Summary"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date, week, driver, route = row[:4]
        if not route:
            continue
        d = parse_date(raw_date)
        if d:
            all_date_routes.add((d, str(route).strip().upper()))

    # Normalize dates to date objects
    normalized = set()
    for d, r in all_date_routes:
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, date):
            pass
        else:
            continue
        normalized.add((d, r))
    all_date_routes = normalized

    print(f"  Found {len(all_date_routes)} unique (date, route) pairs")

    # ================================================================
    # PHASE 2: Create ALL assignments upfront
    # ================================================================
    print("\n=== Phase 2: Creating assignments ===")
    assignment_cache = {}  # (date, route_id) -> assignment_id
    shift_cache = {}  # assignment_id -> shift_id

    sorted_routes = sorted(all_date_routes, key=lambda x: (x[0].isoformat(), x[1]))
    for i, (d, route_upper) in enumerate(sorted_routes):
        if i % 20 == 0:
            print(f"  Processing {i+1}/{len(sorted_routes)}...")
        route_id = route_map_upper.get(route_upper)
        if not route_id:
            continue

        # Find driver for route
        driver_name = None
        for dr_name, dr_routes in driver_route_map.items():
            if route_upper in [r.upper() for r in dr_routes]:
                driver_name = dr_name
                break
        if not driver_name:
            continue

        driver_id = find_driver_id(driver_name)
        v_reg = driver_vehicle_reg.get(driver_name, "")
        vid = vehicle_map.get(v_reg, list(vehicle_map.values())[0] if vehicle_map else None)

        # Find rep for route
        rep_name = route_default_rep.get(route_upper, "NAHASHON NENE")
        rep_id = find_rep_id(rep_name)
        if not rep_id:
            rep_id = list(rep_map.values())[0]

        if not driver_id or not vid:
            continue

        key = (d, route_id)
        if key in assignment_cache:
            continue

        a_id = cuid()
        try:
            cur.execute("""
                INSERT INTO daily_assignments (id, date, "routeId", "salesRepId", "driverId", "vehicleId", status, "createdAt")
                VALUES (%s, %s, %s, %s, %s, %s, 'COMPLETED', NOW())
                ON CONFLICT (date, "routeId") DO NOTHING
            """, (a_id, d, route_id, rep_id, driver_id, vid))
        except:
            pass

        cur.execute('SELECT id FROM daily_assignments WHERE date=%s AND "routeId"=%s', (d, route_id))
        r = cur.fetchone()
        if r:
            assignment_cache[key] = r[0]

            # Create SalesRepShift
            s_id = cuid()
            try:
                cur.execute("""
                    INSERT INTO sales_rep_shifts (id, "assignmentId", "createdAt")
                    VALUES (%s, %s, NOW())
                    ON CONFLICT ("assignmentId") DO NOTHING
                """, (s_id, r[0]))
            except:
                pass

            # Create DriverShift
            ds_id = cuid()
            try:
                cur.execute("""
                    INSERT INTO driver_shifts (id, "assignmentId", "createdAt")
                    VALUES (%s, %s, NOW())
                    ON CONFLICT ("assignmentId") DO NOTHING
                """, (ds_id, r[0]))
            except:
                pass

            cur.execute('SELECT id FROM driver_shifts WHERE "assignmentId"=%s', (r[0],))
            dr = cur.fetchone()
            if dr:
                shift_cache[r[0]] = dr[0]

    cur.execute("SELECT count(*) FROM daily_assignments")
    print(f"  Assignments: {cur.fetchone()[0]}")
    cur.execute("SELECT count(*) FROM driver_shifts")
    print(f"  Driver Shifts: {cur.fetchone()[0]}")

    # ================================================================
    # PHASE 3: Seed Sales → Orders + update SalesRepShifts
    # ================================================================
    print("\n=== Phase 3: Seeding Sales ===")
    ws = wb["Sales breakdown"]
    order_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date, week, rep, route, location, amount, customers = row[:7]
        if not rep or not route:
            continue
        d = parse_date(raw_date)
        if not d:
            continue

        route_id = find_route_id(str(route))
        if not route_id:
            continue

        key = (d, route_id)
        a_id = assignment_cache.get(key)
        if not a_id:
            continue

        amt = pf(amount)
        cust = pi(customers)

        # Update SalesRepShift
        try:
            cur.execute("""
                UPDATE sales_rep_shifts SET
                    "customerCountTarget" = %s, "customerCountActual" = %s,
                    "salesTarget" = %s, "salesActual" = %s
                WHERE "assignmentId" = %s
            """, (cust, cust, amt, amt, a_id))
        except:
            pass

        # Create Order
        try:
            cur.execute("""
                INSERT INTO orders (id, "assignmentId", "customerName", "totalAmount", "createdAt")
                VALUES (%s, %s, %s, %s, NOW())
            """, (cuid(), a_id, f"Daily Sales - {route}", amt))
            order_count += 1
        except:
            pass

    cur.execute("SELECT count(*) FROM orders")
    print(f"  Orders: {cur.fetchone()[0]}")

    # ================================================================
    # PHASE 4: Seed Returns
    # ================================================================
    print("\n=== Phase 4: Seeding Returns ===")
    ws = wb["Returns Summary"]
    returns_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date, week, driver, route, ret_type, product, qty, price, amount, comments = row[:10]
        if not driver or not product:
            continue

        d = parse_date(raw_date)
        if not d:
            continue

        route_name = str(route).strip()
        route_id = find_route_id(route_name)
        if not route_id:
            continue

        key = (d, route_id)
        a_id = assignment_cache.get(key)
        if not a_id:
            continue

        ds_id = shift_cache.get(a_id)
        if not ds_id:
            continue

        sku_id = find_sku_id(product)

        try:
            cur.execute("""
                INSERT INTO returns (id, "driverShiftId", "skuId", type, quantity, price, amount, reason, comments, "createdAt")
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, (cuid(), ds_id, sku_id, normalize_return_type(ret_type),
                  pi(qty), pf(price), pf(amount),
                  f"Product: {product}", str(comments)[:500] if comments else ""))
            returns_count += 1
        except:
            pass

    cur.execute("SELECT count(*) FROM returns")
    print(f"  Returns: {cur.fetchone()[0]} total")

    # ================================================================
    # PHASE 5: Seed Fleet Management
    # ================================================================
    print("\n=== Phase 5: Seeding Fleet Management ===")
    ws = wb["FLEET MANAGEMENT"]

    dates_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    expected_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    actual_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    garage_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    reason_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    tat_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    theft_row = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    theft_reason_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]

    fleet_count = 0
    for col_idx in range(1, len(dates_row)):
        d = parse_date(dates_row[col_idx])
        if not d:
            continue

        expected = pi(expected_row[col_idx]) if col_idx < len(expected_row) else 4
        actual = pi(actual_row[col_idx]) if col_idx < len(actual_row) else 4
        garage_val = str(garage_row[col_idx]).strip() if col_idx < len(garage_row) else "0"
        reason = str(reason_row[col_idx])[:200] if col_idx < len(reason_row) and reason_row[col_idx] else ""
        tat = str(tat_row[col_idx])[:100] if col_idx < len(tat_row) and tat_row[col_idx] else ""
        theft = pi(theft_row[col_idx]) if col_idx < len(theft_row) else 0
        theft_reason = str(theft_reason_row[col_idx])[:200] if col_idx < len(theft_reason_row) and theft_reason_row[col_idx] else ""

        in_garage = 1 if garage_val not in ("0", "N/A", "", "n/a") else 0

        for v_reg, v_id in vehicle_map.items():
            v_garage = 1 if (in_garage and garage_val in v_reg) else 0
            try:
                cur.execute("""
                    INSERT INTO fleet_daily (id, date, "vehicleId", "expectedAvailable", "actualAvailable",
                        "inGarage", "garageReason", "workshopTat", "theftReport", "theftReason",
                        "preDispatchInspection", "createdAt")
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, false, NOW())
                    ON CONFLICT (date, "vehicleId") DO NOTHING
                """, (cuid(), d, v_id, expected, actual, v_garage,
                      reason if v_garage else "", tat if v_garage else "",
                      theft, theft_reason if theft else ""))
                fleet_count += 1
            except:
                pass

    cur.execute("SELECT count(*) FROM fleet_daily")
    print(f"  Fleet Daily: {cur.fetchone()[0]} total")

    # ================================================================
    # PHASE 6: Seed Pricing
    # ================================================================
    print("\n=== Phase 6: Seeding Pricing Surveys ===")
    ws = wb["pricing"]
    default_date = date(2026, 7, 1)

    pricing_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        route_name, rep_name, item, comp_name, comp_price, khel_price, diff = row[:7]
        if not item:
            continue

        route_id = find_route_id(str(route_name)) if route_name else None
        if not route_id:
            continue

        rep_id = find_rep_id(str(rep_name)) if rep_name else list(rep_map.values())[0]
        sku_id = find_sku_id(item)

        try:
            cur.execute("""
                INSERT INTO pricing_surveys (id, date, "routeId", "salesRepId", "skuId",
                    "competitorName", "competitorPrice", "khelPrice", difference, "createdAt")
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, (cuid(), default_date, route_id, rep_id, sku_id,
                  str(comp_name)[:100] if comp_name else "", pf(comp_price), pf(khel_price),
                  pf(diff) if diff else pf(khel_price) - pf(comp_price)))
            pricing_count += 1
        except:
            pass

    cur.execute("SELECT count(*) FROM pricing_surveys")
    print(f"  Pricing Surveys: {cur.fetchone()[0]} total")

    # ================================================================
    # PHASE 7: Seed Challenges
    # ================================================================
    print("\n=== Phase 7: Seeding Challenges ===")
    ws = wb["CHALLENGES"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        no, gap, what, who, when_val, resolved = row[:6]
        if not gap:
            continue

        when_d = parse_date(when_val)
        resolved_bool = str(resolved).strip().upper() == "YES" if resolved else False

        try:
            cur.execute("""
                INSERT INTO challenges (id, date, gap, "whatAction", who, "when", resolved, "createdAt")
                VALUES (%s, NOW(), %s, %s, %s, %s, %s, NOW())
            """, (cuid(), str(gap)[:500], str(what)[:500] if what else "",
                  str(who)[:200] if who else "", when_d, resolved_bool))
        except:
            pass

    cur.execute("SELECT count(*) FROM challenges")
    print(f"  Challenges: {cur.fetchone()[0]} total")

    # ================================================================
    # FINAL SUMMARY
    # ================================================================
    print("\n=== FINAL COUNTS ===")
    for table in ["daily_assignments", "sales_rep_shifts", "driver_shifts", "orders",
                   "returns", "fleet_daily", "pricing_surveys", "challenges"]:
        cur.execute(f'SELECT count(*) FROM "{table}"')
        print(f"  {table}: {cur.fetchone()[0]}")

    cur.close()
    conn.close()
    print("\nAll Excel data seeded into Neon!")


if __name__ == "__main__":
    main()
